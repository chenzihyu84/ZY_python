from random import randint,choices
import csv
from csv import DictWriter
from openpyxl import Workbook

def getStudents(nums:int) -> list[dict]:
    
 
    students:list[dict] = []
    with open('names.txt', mode='r',encoding = 'utf-8') as file:
        names:str = file.read()
    namesList:list[str] = names.split('\n')
    names = choices(namesList,k=nums)
    
    for i in range(nums):
        stu = {
        '姓名':names[i],
        '國文':randint(45,100),
        '數學':randint(45,100),
        '英文':randint(45,100),
        '地理':randint(45,100),
        }
        students.append(stu)
    return students

def save_to_csv(students:list[dict],fileName:str)->None:
    fileNameWithExtension:list = fileName + '.csv'
    with open(fileNameWithExtension,mode = 'w',encoding = 'utf-8',newline = '') as file:
        fieldnames:list[str] = ['姓名','國文','數學','英文','地理']
        writer:DictWriter= csv.DictWriter(file,fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(students)
    print("寫入成功")


def save_to_excel(students:list[dict],fileName:str) -> None:
    fileNameWithExtension:str = fileName + ".xlsx"
    #print(f"檔案名是{fileNameWithExtension}")

    wb = Workbook()
    ws = wb.active
    ws.title = fileName
    ws['A1'] = "姓名"
    ws['B1'] = "國文"
    ws['C1'] = "數學"
    ws['D1'] = "英文"
    ws['E1'] = "地理"
    for student in students:
        #ws.append(['昌嘉洋',81,65,65,86])
        studentData:list = list(student.values())
        ws.append(studentData)
    wb.save(fileNameWithExtension)

   